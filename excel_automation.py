import openpyxl
from openpyxl.utils import *
import os
import xlwings
# print(os.getcwd())
wb = openpyxl.load_workbook('debtors.xlsx')
# print(wb.get_sheet_names())
print(wb.sheetnames)

sheet = wb['CustAging_RO.Report']
print(sheet['A1'].value)

variable = sheet['B1']
print(variable.value)
print(variable.row)
print(variable.column)
print(get_column_letter(2))
print(variable.coordinate)

print(sheet.cell(row=1, column=2).value)

for i in range(1,4):
    print(i, sheet.cell(row=i, column=2).value)





